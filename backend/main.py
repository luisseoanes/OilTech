# main.py

from typing import List
import os, uuid, re
from dotenv import load_dotenv
load_dotenv()
from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import FileResponse, Response, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from datetime import timedelta
from fastapi.security import OAuth2PasswordRequestForm
import models, schemas, database, auth

# Create tables
models.Base.metadata.create_all(bind=database.engine)

def apply_migrations():
    import sqlite3
    try:
        # Get path from SQLAlchemy engine
        path = str(database.engine.url).replace("sqlite:///", "")
        if not path or path == "sqlite://": path = "oiltech.db"
        
        conn = sqlite3.connect(path)
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(quotations)")
        columns = [col[1] for col in cursor.fetchall()]
        if "reference" not in columns:
            print("Migrating: Adding 'reference' column to quotations")
            cursor.execute("ALTER TABLE quotations ADD COLUMN reference TEXT")
            conn.commit()

        cursor.execute("PRAGMA table_info(products)")
        product_columns = [col[1] for col in cursor.fetchall()]
        if "subcategory" not in product_columns:
            print("Migrating: Adding 'subcategory' column to products")
            cursor.execute("ALTER TABLE products ADD COLUMN subcategory TEXT")
            conn.commit()
        if "price" in product_columns:
            print("Migrating: Dropping 'price' column from products")
            cursor.execute("ALTER TABLE products DROP COLUMN price")
            conn.commit()
        if "code" in product_columns:
            print("Migrating: Dropping 'code' column from products")
            cursor.execute("ALTER TABLE products DROP COLUMN code")
            conn.commit()
        if "options" in product_columns:
            print("Migrating: Dropping 'options' column from products")
            cursor.execute("ALTER TABLE products DROP COLUMN options")
            conn.commit()
        if "category" in product_columns:
            print("Migrating: Dropping 'category' column from products")
            cursor.execute("ALTER TABLE products DROP COLUMN category")
            conn.commit()
        if "category_id" not in product_columns:
            print("Migrating: Adding 'category_id' column to products")
            cursor.execute("ALTER TABLE products ADD COLUMN category_id INTEGER REFERENCES categories(id)")
            conn.commit()
        if "price_text" not in product_columns:
            print("Migrating: Adding 'price_text' column to products")
            cursor.execute("ALTER TABLE products ADD COLUMN price_text TEXT")
            conn.commit()
        if "brands" not in product_columns:
            print("Migrating: Adding 'brands' column to products")
            cursor.execute("ALTER TABLE products ADD COLUMN brands TEXT")
            conn.commit()
        if "subcategory" in product_columns:
            print("Migrating: Dropping 'subcategory' TEXT column from products")
            cursor.execute("ALTER TABLE products DROP COLUMN subcategory")
            conn.commit()
        if "subcategory_id" not in product_columns:
            print("Migrating: Adding 'subcategory_id' column to products")
            cursor.execute("ALTER TABLE products ADD COLUMN subcategory_id INTEGER REFERENCES subcategories(id)")
            conn.commit()

        cursor.execute("PRAGMA table_info(quotations)")
        quotation_columns = [col[1] for col in cursor.fetchall()]
        if "total_estimated" in quotation_columns:
            print("Migrating: Dropping 'total_estimated' column from quotations")
            cursor.execute("ALTER TABLE quotations DROP COLUMN total_estimated")
            conn.commit()

        # Re-fetch product columns after potential drops above
        cursor.execute("PRAGMA table_info(products)")
        product_columns = [col[1] for col in cursor.fetchall()]
        if "brands" in product_columns:
            print("Migrating: Dropping legacy 'brands' text column from products")
            cursor.execute("ALTER TABLE products DROP COLUMN brands")
            conn.commit()

        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='brands'")
        if not cursor.fetchone():
            print("Migrating: Creating 'brands' table")
            cursor.execute("""
                CREATE TABLE brands (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT UNIQUE NOT NULL,
                    image_url TEXT
                )
            """)
            conn.commit()

        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='product_brands'")
        if not cursor.fetchone():
            print("Migrating: Creating 'product_brands' junction table")
            cursor.execute("""
                CREATE TABLE product_brands (
                    product_id INTEGER NOT NULL REFERENCES products(id),
                    brand_id INTEGER NOT NULL REFERENCES brands(id),
                    PRIMARY KEY (product_id, brand_id)
                )
            """)
            conn.commit()

        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='presentations'")
        if not cursor.fetchone():
            print("Migrating: Creating 'presentations' table")
            cursor.execute("""
                CREATE TABLE presentations (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT UNIQUE NOT NULL
                )
            """)
            conn.commit()

        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='product_presentations'")
        if not cursor.fetchone():
            print("Migrating: Creating 'product_presentations' junction table")
            cursor.execute("""
                CREATE TABLE product_presentations (
                    product_id INTEGER NOT NULL REFERENCES products(id),
                    presentation_id INTEGER NOT NULL REFERENCES presentations(id),
                    PRIMARY KEY (product_id, presentation_id)
                )
            """)
            conn.commit()

        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='sales'")
        if not cursor.fetchone():
            print("Migrating: Creating 'sales' table")
            cursor.execute("""
                CREATE TABLE sales (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    quotation_id INTEGER NOT NULL UNIQUE REFERENCES quotations(id),
                    price REAL NOT NULL,
                    items TEXT,
                    customer_name TEXT NOT NULL DEFAULT '',
                    customer_contact TEXT NOT NULL DEFAULT '',
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """)
            conn.commit()
        else:
            cursor.execute("PRAGMA table_info(sales)")
            sales_columns = [col[1] for col in cursor.fetchall()]
            if "customer_name" not in sales_columns:
                print("Migrating: Adding 'customer_name' to sales")
                cursor.execute("ALTER TABLE sales ADD COLUMN customer_name TEXT NOT NULL DEFAULT ''")
                conn.commit()
            if "customer_contact" not in sales_columns:
                print("Migrating: Adding 'customer_contact' to sales")
                cursor.execute("ALTER TABLE sales ADD COLUMN customer_contact TEXT NOT NULL DEFAULT ''")
                conn.commit()

        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='site_assets'")
        if not cursor.fetchone():
            print("Migrating: Creating 'site_assets' table")
            cursor.execute("""
                CREATE TABLE site_assets (
                    key TEXT PRIMARY KEY,
                    description TEXT,
                    image_url TEXT,
                    display_mode TEXT NOT NULL DEFAULT 'single',
                    gallery_urls TEXT,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """)
            conn.commit()
        else:
            cursor.execute("PRAGMA table_info(site_assets)")
            site_asset_columns = [col[1] for col in cursor.fetchall()]
            if "display_mode" not in site_asset_columns:
                print("Migrating: Adding 'display_mode' to site_assets")
                cursor.execute("ALTER TABLE site_assets ADD COLUMN display_mode TEXT NOT NULL DEFAULT 'single'")
                conn.commit()
            if "gallery_urls" not in site_asset_columns:
                print("Migrating: Adding 'gallery_urls' to site_assets")
                cursor.execute("ALTER TABLE site_assets ADD COLUMN gallery_urls TEXT")
                conn.commit()

        conn.close()
    except Exception as e:
        print(f"Migration error: {e}")

apply_migrations()

# Auto-seed admin user if it doesn't exist
def seed_admin():
    db = database.SessionLocal()
    try:
        admin = db.query(models.User).filter(models.User.username == "admin").first()
        if not admin:
            hashed_pw = auth.get_password_hash("admin123")
            admin_user = models.User(username="admin", hashed_password=hashed_pw)
            db.add(admin_user)
            db.commit()
            print("Auto-seeded admin user")
    finally:
        db.close()

seed_admin()

# Auto-seed initial categories
def seed_categories():
    db = database.SessionLocal()
    try:
        count = db.query(models.Category).count()
        if count == 0:
            initial = [
                {"name": "Lubricantes Industriales", "tags": "lubricantes aceites hidraulico ISO32 ISO46 ISO68 ISO100 compresores engranajes sintetico termico dielectrico reductores corte neumatico"},
                {"name": "Grasas Industriales", "tags": "grasas rodamientos altapresion maquinariapesada lubricacioncentralizada"},
                {"name": "Seguridad Industrial (EPP)", "tags": "EPP seguridad guantes botas arnes eslingas proteccionauditiva gafas overol chalecoreflectivo careta"},
                {"name": "Productos de Limpieza y Mantenimiento", "tags": "limpieza mantenimiento alcoholindustrial amoniocu jabones limpiavidrios"},
                {"name": "Herramientas y Suministros Técnicos", "tags": "herramientas suministros bombas acoples bandas filtros componenteselectricos"}
            ]
            for cat in initial:
                db_cat = models.Category(**cat)
                db.add(db_cat)
            db.commit()
            print("Auto-seeded categories")
    finally:
        db.close()

seed_categories()

def seed_subcategories():
    db = database.SessionLocal()
    try:
        if db.query(models.Subcategory).count() > 0:
            return
        subcategories = {
            "Lubricantes Industriales": [
                "Aceite hidráulico ISO 32", "Aceite hidráulico ISO 46",
                "Aceite hidráulico ISO 68", "Aceite hidráulico ISO 100",
                "Aceites para compresores", "Aceites para engranajes industriales",
                "Aceites sintéticos", "Aceites térmicos", "Aceites dieléctricos",
                "Aceites para reductores", "Aceites de corte", "Aceites solubles",
                "Aceites para mantenimiento neumático"
            ],
            "Grasas Industriales": [
                "Rodamientos", "Equipos industriales",
                "Maquinaria pesada", "Sistemas de lubricación centralizada"
            ],
            "Seguridad Industrial (EPP)": [
                "Guantes industriales", "Botas de seguridad", "Arnés de seguridad",
                "Eslingas y líneas de vida", "Protectores auditivos", "Gafas de seguridad",
                "Overoles industriales", "Chalecos reflectivos", "Caretas de protección"
            ],
            "Productos de Limpieza y Mantenimiento": [
                "Alcohol industrial", "Amonio cuaternario",
                "Jabones industriales", "Limpiavidrios"
            ],
            "Herramientas y Suministros Técnicos": [
                "Bombas industriales", "Acoples y conexiones", "Boquillas",
                "Bandas industriales", "Filtros", "Abrazaderas", "Arandelas",
                "Herramientas manuales", "Componentes eléctricos", "Equipos para mantenimiento"
            ]
        }
        for cat_name, subcat_names in subcategories.items():
            cat = db.query(models.Category).filter(models.Category.name == cat_name).first()
            if not cat:
                continue
            for name in subcat_names:
                db.add(models.Subcategory(name=name, category_id=cat.id))
        db.commit()
        print("Auto-seeded subcategories")
    finally:
        db.close()

seed_subcategories()

def seed_brands():
    db = database.SessionLocal()
    try:
        if db.query(models.Brand).count() > 0:
            return
        initial = [
            {"name": "Shell",     "image_url": "https://images.icon-icons.com/2699/PNG/512/shell_logo_icon_168832.png"},
            {"name": "Mobil",     "image_url": "https://upload.wikimedia.org/wikipedia/commons/thumb/3/3d/Mobil_logo.svg/3840px-Mobil_logo.svg.png"},
            {"name": "Terpel",    "image_url": "https://portalcolombia.terpel.com/static/images/terpel_logo_og.png"},
            {"name": "Chevron",   "image_url": "https://upload.wikimedia.org/wikipedia/commons/thumb/8/86/Chevron_Logo.svg/960px-Chevron_Logo.svg.png"},
            {"name": "Repsol",    "image_url": "https://upload.wikimedia.org/wikipedia/commons/thumb/e/e8/Repsol_logo.svg/1280px-Repsol_logo.svg.png"},
            {"name": "3M",        "image_url": "https://upload.wikimedia.org/wikipedia/commons/thumb/1/15/3M_wordmark.svg/3840px-3M_wordmark.svg.png"},
            {"name": "Castrol",   "image_url": "https://upload.wikimedia.org/wikipedia/commons/thumb/4/40/Castrol_logo_2023.svg/3840px-Castrol_logo_2023.svg.png"},
            {"name": "Valvoline", "image_url": "https://upload.wikimedia.org/wikipedia/commons/thumb/3/34/Valvoline_company_logo.svg/1280px-Valvoline_company_logo.svg.png"},
            {"name": "Gulf",      "image_url": "https://upload.wikimedia.org/wikipedia/commons/7/70/Gulf_logo.png"},
            {"name": "Motul",     "image_url": "https://1000marcas.net/wp-content/uploads/2021/05/Logo-Motul.png"},
            {"name": "MCR Safety","image_url": "https://ibtinc.com/wp-content/uploads/2022/12/MCR-Safety-logo.png"},
            {"name": "Ansell",    "image_url": "https://upload.wikimedia.org/wikipedia/en/thumb/0/07/Ansell_logo.svg/3840px-Ansell_logo.svg.png"},
            {"name": "Bosch",     "image_url": "https://upload.wikimedia.org/wikipedia/commons/c/c3/Bosch_logo.png"},
            {"name": "ARO",       "image_url": "https://sumicali.com/cdn/shop/collections/ARO-Sumicali.png?crop=center&height=2048&v=1716414249&width=2048"},
        ]
        for b in initial:
            db.add(models.Brand(**b))
        db.commit()
        print("Auto-seeded brands")
    finally:
        db.close()

seed_brands()

def seed_site_assets():
    db = database.SessionLocal()
    try:
        initial = [
            {"key": "logo_navbar", "description": "Logo de la barra de navegación", "image_url": "imagenes/LogoNavBar.png"},
            {"key": "hero_bg", "description": "Imagen de fondo del Hero (Inicio)", "image_url": "imagenes/hero_industrial.png"},
            {"key": "cat_lubricantes", "description": "Imagen: Línea Lubricantes Industriales", "image_url": "imagenes/cat_lubricantes.png"},
            {"key": "cat_grasas", "description": "Imagen: Línea Grasas Industriales", "image_url": "imagenes/GrasaIndustrial.png"},
            {"key": "cat_seguridad", "description": "Imagen: Línea Seguridad Industrial (EPP)", "image_url": "imagenes/SeguridadIndustrial.avif"},
            {"key": "cat_limpieza", "description": "Imagen: Línea Limpieza y Mantenimiento", "image_url": "imagenes/LimpiadoresIndustriales.png"},
            {"key": "cat_herramientas", "description": "Imagen: Línea Herramientas Técnicas", "image_url": "imagenes/HerramientasTecnicas.png"},
            {"key": "benefit_1", "description": "Icono: Equipo Técnico Especializado", "image_url": "imagenes/svg01.png"},
            {"key": "benefit_2", "description": "Icono: Portafolio Integral", "image_url": "imagenes/svg02.png"},
            {"key": "benefit_3", "description": "Icono: Cobertura Nacional", "image_url": "imagenes/svg03.png"},
            {"key": "benefit_4", "description": "Icono: Calidad Garantizada", "image_url": "imagenes/svg04.png"},
            {"key": "line_lubricantes", "description": "Imagen Detalle: Lubricantes Industriales", "image_url": "imagenes/oil1.png"},
            {"key": "line_grasas", "description": "Imagen Detalle: Grasas Industriales", "image_url": "imagenes/oil2.png"},
            {"key": "line_seguridad", "description": "Imagen Detalle: Seguridad Industrial", "image_url": "imagenes/oil3.png"},
            {"key": "line_herramientas", "description": "Imagen Detalle: Herramientas Técnicas", "image_url": "imagenes/oil4.png"},
            {"key": "badge_colombia", "description": "Sello: Sello Colombia", "image_url": "imagenes/SelloColombia.png"},
            {"key": "badge_sostenibilidad", "description": "Sello: Sostenibilidad", "image_url": "imagenes/SelloSostenibilidad.png"},
            {"key": "badge_calidad", "description": "Sello: Calidad Técnica", "image_url": "imagenes/SelloCalidad.png"},
            {"key": "footer_logo", "description": "Logo del pie de página", "image_url": "imagenes/LogoHero.png"},
            {"key": "logo_iso", "description": "Certificación: Logo ISO", "image_url": "imagenes/LogoISO.png"},
            {"key": "logo_icontec", "description": "Certificación: Logo Icontec", "image_url": "imagenes/LogoIcontec.png"},
        ]
        for asset in initial:
            if not db.query(models.SiteAsset).filter(models.SiteAsset.key == asset["key"]).first():
                db.add(models.SiteAsset(**asset))
        db.commit()
        print("Auto-seeded site assets")
    finally:
        db.close()

seed_site_assets()

app = FastAPI()

# CORS - Allow explicit origins for Vercel and local development
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5500",
        "http://127.0.0.1:5500",
        "https://oil-tech.vercel.app",
        "https://oiltechdecolombiasas.com",
        "https://www.oiltechdecolombiasas.com"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.middleware("http")
async def add_pna_header(request, call_next):
    response = await call_next(request)
    response.headers["Access-Control-Allow-Private-Network"] = "true"
    return response

# Dependency
def get_db():
    db = database.SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Auth Endpoints
@app.post("/token", response_model=schemas.Token)
async def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.username == form_data.username).first()
    if not user or not auth.verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=auth.ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = auth.create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/users/me", response_model=schemas.UserCreate) # Only returning username for now
async def read_users_me(current_user: models.User = Depends(auth.get_current_user)):
    return {"username": current_user.username, "password": ""} # Don't return password hash

# Category Endpoints
@app.get("/categories/", response_model=List[schemas.Category])
def read_categories(db: Session = Depends(get_db)):
    return db.query(models.Category).all()

@app.post("/categories/", response_model=schemas.Category)
def create_category(category: schemas.CategoryCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    db_cat = models.Category(**category.dict())
    db.add(db_cat)
    db.commit()
    db.refresh(db_cat)
    return db_cat

@app.put("/categories/{category_id}", response_model=schemas.Category)
def update_category(category_id: int, category: schemas.CategoryCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    db_cat = db.query(models.Category).filter(models.Category.id == category_id).first()
    if not db_cat:
        raise HTTPException(status_code=404, detail="Category not found")
    db_cat.name = category.name
    db_cat.tags = category.tags
    db.commit()
    db.refresh(db_cat)
    return db_cat

@app.delete("/categories/{category_id}")
def delete_category(category_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    db_cat = db.query(models.Category).filter(models.Category.id == category_id).first()
    if not db_cat:
        raise HTTPException(status_code=404, detail="Category not found")
    db.query(models.Product).filter(models.Product.category_id == category_id).update({"category_id": None, "subcategory_id": None})
    db.query(models.Subcategory).filter(models.Subcategory.category_id == category_id).delete()
    db.delete(db_cat)
    db.commit()
    return {"ok": True}

# Subcategory Endpoints
@app.get("/subcategories/", response_model=List[schemas.Subcategory])
def read_subcategories(category_id: int = None, db: Session = Depends(get_db)):
    query = db.query(models.Subcategory)
    if category_id:
        query = query.filter(models.Subcategory.category_id == category_id)
    return query.all()

@app.post("/subcategories/", response_model=schemas.Subcategory)
def create_subcategory(subcategory: schemas.SubcategoryCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    db_subcat = models.Subcategory(**subcategory.dict())
    db.add(db_subcat)
    db.commit()
    db.refresh(db_subcat)
    return db_subcat

@app.put("/subcategories/{subcategory_id}", response_model=schemas.Subcategory)
def update_subcategory(subcategory_id: int, subcategory: schemas.SubcategoryCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    db_subcat = db.query(models.Subcategory).filter(models.Subcategory.id == subcategory_id).first()
    if not db_subcat:
        raise HTTPException(status_code=404, detail="Subcategory not found")
    db_subcat.name = subcategory.name
    db_subcat.category_id = subcategory.category_id
    db.commit()
    db.refresh(db_subcat)
    return db_subcat

@app.delete("/subcategories/{subcategory_id}")
def delete_subcategory(subcategory_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    db_subcat = db.query(models.Subcategory).filter(models.Subcategory.id == subcategory_id).first()
    if not db_subcat:
        raise HTTPException(status_code=404, detail="Subcategory not found")
    db.query(models.Product).filter(models.Product.subcategory_id == subcategory_id).update({"subcategory_id": None})
    db.delete(db_subcat)
    db.commit()
    return {"ok": True}

# Brand Endpoints
@app.get("/brands/", response_model=List[schemas.Brand])
def read_brands(db: Session = Depends(get_db)):
    return db.query(models.Brand).all()

@app.post("/brands/", response_model=schemas.Brand)
def create_brand(brand: schemas.BrandCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    db_brand = models.Brand(**brand.dict())
    db.add(db_brand)
    db.commit()
    db.refresh(db_brand)
    return db_brand

@app.put("/brands/{brand_id}", response_model=schemas.Brand)
def update_brand(brand_id: int, brand: schemas.BrandCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    db_brand = db.query(models.Brand).filter(models.Brand.id == brand_id).first()
    if not db_brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    db_brand.name = brand.name
    db_brand.image_url = brand.image_url
    db.commit()
    db.refresh(db_brand)
    return db_brand

@app.delete("/brands/{brand_id}")
def delete_brand(brand_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    db_brand = db.query(models.Brand).filter(models.Brand.id == brand_id).first()
    if not db_brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    db.delete(db_brand)
    db.commit()
    return {"ok": True}

# Presentation Endpoints
@app.get("/presentations/", response_model=List[schemas.Presentation])
def read_presentations(db: Session = Depends(get_db)):
    return db.query(models.Presentation).all()

@app.post("/presentations/", response_model=schemas.Presentation)
def create_presentation(presentation: schemas.PresentationCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    db_pres = models.Presentation(**presentation.dict())
    db.add(db_pres)
    db.commit()
    db.refresh(db_pres)
    return db_pres

@app.put("/presentations/{presentation_id}", response_model=schemas.Presentation)
def update_presentation(presentation_id: int, presentation: schemas.PresentationCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    db_pres = db.query(models.Presentation).filter(models.Presentation.id == presentation_id).first()
    if not db_pres:
        raise HTTPException(status_code=404, detail="Presentation not found")
    db_pres.name = presentation.name
    db.commit()
    db.refresh(db_pres)
    return db_pres

@app.delete("/presentations/{presentation_id}")
def delete_presentation(presentation_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    db_pres = db.query(models.Presentation).filter(models.Presentation.id == presentation_id).first()
    if not db_pres:
        raise HTTPException(status_code=404, detail="Presentation not found")
    db.delete(db_pres)
    db.commit()
    return {"ok": True}

# Product Endpoints
@app.get("/products/", response_model=List[schemas.Product])
def read_products(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    products = db.query(models.Product).offset(skip).limit(limit).all()
    return products

@app.post("/products/", response_model=schemas.Product)
def create_product(product: schemas.ProductCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    data = product.dict()
    brand_ids = data.pop("brand_ids", []) or []
    presentation_ids = data.pop("presentation_ids", []) or []
    data["search_tags"] = ",".join(data["name"].split())
    db_product = models.Product(**data)
    if brand_ids:
        db_product.brands = db.query(models.Brand).filter(models.Brand.id.in_(brand_ids)).all()
    if presentation_ids:
        db_product.presentations = db.query(models.Presentation).filter(models.Presentation.id.in_(presentation_ids)).all()
    db.add(db_product)
    db.commit()
    db.refresh(db_product)
    return db_product

@app.put("/products/{product_id}", response_model=schemas.Product)
def update_product(product_id: int, product: schemas.ProductCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    db_product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if db_product is None:
        raise HTTPException(status_code=404, detail="Product not found")

    data = product.dict()
    brand_ids = data.pop("brand_ids", []) or []
    presentation_ids = data.pop("presentation_ids", []) or []
    data["search_tags"] = ",".join(data["name"].split())
    for key, value in data.items():
        setattr(db_product, key, value)
    db_product.brands = db.query(models.Brand).filter(models.Brand.id.in_(brand_ids)).all()
    db_product.presentations = db.query(models.Presentation).filter(models.Presentation.id.in_(presentation_ids)).all()

    db.commit()
    db.refresh(db_product)
    return db_product

@app.delete("/products/{product_id}")
def delete_product(product_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    db_product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if db_product is None:
        raise HTTPException(status_code=404, detail="Product not found")

    db.delete(db_product)
    db.commit()
    return {"ok": True}

@app.get("/products/import-template")
def download_import_template(current_user: models.User = Depends(auth.get_current_user)):
    import openpyxl, io
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(["nombre", "categoria", "subcategoria", "marcas", "presentaciones", "descripcion", "imagen_url"])
    ws.append(["Aceite Hidráulico ISO 46", "Lubricantes Industriales", "Aceite hidráulico ISO 46", "Shell,Mobil", "1 Galón,1 Cuñete", "Aceite hidráulico de alta calidad", ""])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 30
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_productos.xlsx"}
    )

@app.post("/products/preview-excel")
async def preview_excel_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    import openpyxl, io
    if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx o .xlsm")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Archivo Excel inválido o corrupto")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="El archivo no contiene datos (mínimo encabezado + 1 fila)")
    categories = {c.name.strip().lower(): c for c in db.query(models.Category).all()}
    all_subcats = db.query(models.Subcategory).all()
    brands = {b.name.strip().lower(): b for b in db.query(models.Brand).all()}
    presentations = {p.name.strip().lower(): p for p in db.query(models.Presentation).all()}
    results = []
    for row_num, row in enumerate(rows[1:], start=2):
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue
        def cell(idx):
            val = row[idx] if idx < len(row) else None
            return str(val).strip() if val is not None else ''
        nombre = cell(0)
        categoria_str = cell(1)
        subcategoria_str = cell(2)
        marcas_str = cell(3)
        presentaciones_str = cell(4)
        descripcion = cell(5)
        imagen_url = cell(6)
        warnings, errors = [], []
        if not nombre:
            errors.append("Nombre vacío")
        cat = categories.get(categoria_str.lower()) if categoria_str else None
        if not cat:
            if categoria_str:
                warnings.append(f"Categoría '{categoria_str}' no encontrada")
            errors.append("Categoría requerida")
        subcat = None
        if subcategoria_str:
            for sc in all_subcats:
                if sc.name.strip().lower() == subcategoria_str.lower() and cat and sc.category_id == cat.id:
                    subcat = sc
                    break
            if not subcat:
                for sc in all_subcats:
                    if sc.name.strip().lower() == subcategoria_str.lower():
                        subcat = sc
                        break
            if not subcat:
                warnings.append(f"Subcategoría '{subcategoria_str}' no encontrada")
        brand_ids, brand_names = [], []
        for b_name in [x.strip() for x in marcas_str.split(',') if x.strip()]:
            b = brands.get(b_name.lower())
            if b:
                brand_ids.append(b.id)
                brand_names.append(b.name)
            else:
                warnings.append(f"Marca '{b_name}' no encontrada")
        pres_ids, pres_names = [], []
        for p_name in [x.strip() for x in presentaciones_str.split(',') if x.strip()]:
            p = presentations.get(p_name.lower())
            if p:
                pres_ids.append(p.id)
                pres_names.append(p.name)
            else:
                warnings.append(f"Presentación '{p_name}' no encontrada")
        results.append({
            "row": row_num,
            "name": nombre,
            "category_id": cat.id if cat else None,
            "category_name": cat.name if cat else categoria_str,
            "subcategory_id": subcat.id if subcat else None,
            "subcategory_name": subcat.name if subcat else subcategoria_str,
            "brand_ids": brand_ids,
            "brand_names": brand_names,
            "presentation_ids": pres_ids,
            "presentation_names": pres_names,
            "description": descripcion or None,
            "image_url": imagen_url or None,
            "warnings": warnings,
            "errors": errors,
            "status": "error" if errors else ("warning" if warnings else "ok")
        })
    return results

@app.post("/products/bulk-import")
def bulk_import_products(
    products: List[schemas.ProductCreate],
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    created, errors = 0, []
    for i, product in enumerate(products):
        try:
            data = product.dict()
            brand_ids = data.pop("brand_ids", []) or []
            presentation_ids = data.pop("presentation_ids", []) or []
            data["search_tags"] = ",".join(data["name"].split())
            db_product = models.Product(**data)
            if brand_ids:
                db_product.brands = db.query(models.Brand).filter(models.Brand.id.in_(brand_ids)).all()
            if presentation_ids:
                db_product.presentations = db.query(models.Presentation).filter(models.Presentation.id.in_(presentation_ids)).all()
            db.add(db_product)
            db.commit()
            db.refresh(db_product)
            created += 1
        except Exception as e:
            db.rollback()
            errors.append({"index": i, "error": str(e)})
    return {"created": created, "errors": errors}

# Sale Endpoints
@app.get("/sales/", response_model=List[schemas.Sale])
def read_sales(db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    return db.query(models.Sale).order_by(models.Sale.created_at.desc()).all()

@app.post("/sales/", response_model=schemas.Sale)
def create_sale(sale: schemas.SaleCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    quotation = db.query(models.Quotation).filter(models.Quotation.id == sale.quotation_id).first()
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")
    existing = db.query(models.Sale).filter(models.Sale.quotation_id == sale.quotation_id).first()
    if existing:
        raise HTTPException(status_code=400, detail="This quotation already has a sale")
    db_sale = models.Sale(
        quotation_id=sale.quotation_id,
        price=sale.price,
        items=[item.dict() for item in sale.items],
        customer_name=sale.customer_name,
        customer_contact=sale.customer_contact,
    )
    db.add(db_sale)
    quotation.status = "Purchased"
    db.commit()
    db.refresh(db_sale)
    return db_sale

# Quotation Endpoints
@app.post("/quotations/", response_model=schemas.Quotation)
def create_quotation(quotation: schemas.QuotationCreate, db: Session = Depends(get_db)):
    db_quotation = models.Quotation(**quotation.dict())
    db.add(db_quotation)
    db.commit()
    db.refresh(db_quotation)
    db_quotation.reference = f"COT-{db_quotation.id:06d}"
    db.commit()
    db.refresh(db_quotation)
    return db_quotation

@app.get("/stats", response_model=schemas.AnalyticsStats)
def get_stats(db: Session = Depends(get_db)):
    quotations = db.query(models.Quotation).all()
    
    total_quoted = db.query(models.Quotation).count()
    total_purchased = db.query(models.Quotation).filter(models.Quotation.status == "Purchased").count()
    
    # Calculate top products
    product_counts = {}
    for q in quotations:
        if q.items:
            # Items are stored as JSON list of dicts
            for item in q.items:
                name = item.get("product_name", "Unknown")
                qty = item.get("quantity", 1)
                product_counts[name] = product_counts.get(name, 0) + qty
    
    top_products = [
        schemas.TopProduct(name=name, count=count)
        for name, count in sorted(product_counts.items(), key=lambda item: item[1], reverse=True)[:5]
    ]

    # Calculate sales history (purchased only — count per day)
    sales_by_date = {}
    for q in quotations:
        if q.status == "Purchased":
            date_str = q.created_at.strftime("%Y-%m-%d")
            sales_by_date[date_str] = sales_by_date.get(date_str, 0) + 1

    sales_history = [
        schemas.SalesData(date=date, amount=float(count))
        for date, count in sorted(sales_by_date.items())
    ]
    
    return schemas.AnalyticsStats(
        total_quoted=total_quoted,
        total_purchased=total_purchased,
        top_products=top_products,
        sales_history=sales_history
    )

@app.get("/quotations/", response_model=List[schemas.Quotation])
def read_quotations(skip: int = 0, limit: int = 100, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    quotations = db.query(models.Quotation).order_by(models.Quotation.created_at.desc()).offset(skip).limit(limit).all()
    return quotations

@app.put("/quotations/{quotation_id}/status")
def update_quotation_status(quotation_id: int, status: str, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    db_quotation = db.query(models.Quotation).filter(models.Quotation.id == quotation_id).first()
    if db_quotation is None:
        raise HTTPException(status_code=404, detail="Quotation not found")
    
    db_quotation.status = status
    db.commit()
    return {"ok": True}



import shutil

PDF_DIR = os.path.join(database.VOLUME_PATH, "fichas-tecnicas")
os.makedirs(PDF_DIR, exist_ok=True)

_PDF_FILENAME_RE = re.compile(r'^[a-f0-9]{32}\.pdf$')

@app.post("/upload/pdf")
async def upload_pdf(
    file: UploadFile = File(...),
    current_user: models.User = Depends(auth.get_current_user)
):
    content = await file.read()

    # Verificar magic bytes — los PDF reales empiezan con %PDF
    if not content.startswith(b"%PDF"):
        raise HTTPException(status_code=400, detail="El archivo no es un PDF válido")

    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="El PDF no puede superar 20MB")

    filename = f"{uuid.uuid4().hex}.pdf"
    filepath = os.path.join(PDF_DIR, filename)
    with open(filepath, "wb") as f:
        f.write(content)

    return {"url": f"/files/pdf/{filename}"}


@app.get("/files/pdf/{filename}")
async def serve_pdf(filename: str):
    # Validar que el nombre sea exactamente el patrón que nosotros generamos
    # Esto elimina path traversal: ../.. u otros caracteres no pueden pasar el regex
    if not _PDF_FILENAME_RE.match(filename):
        raise HTTPException(status_code=400, detail="Nombre de archivo inválido")

    filepath = os.path.join(PDF_DIR, filename)
    if not os.path.isfile(filepath):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    with open(filepath, "rb") as f:
        content = f.read()

    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": "inline; filename=\"ficha-tecnica.pdf\""},
    )


# ... existing imports ...

@app.post("/admin/upload-db")
async def upload_db(file: UploadFile = File(...), current_user: models.User = Depends(auth.get_current_user)):
    """
    Endpoint temporal para subir la base de datos sqlite al volumen de Railway.
    Sobreescribe la base de datos actual.
    """
    try:
        # Use the path defined in database.py
        destination_path = database.DB_URL_PATH
        
        with open(destination_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        return {"filename": file.filename, "message": "Database uploaded successfully to " + destination_path}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not upload database: {str(e)}")

@app.get("/admin/download-db")
async def download_db(current_user: models.User = Depends(auth.get_current_user)):
    """
    Endpoint temporal para descargar la base de datos sqlite del volumen de Railway.
    """
    try:
        db_path = database.DB_URL_PATH
        if not db_path or not os.path.exists(db_path):
            raise HTTPException(status_code=404, detail="Database file not found")
        return FileResponse(
            path=db_path,
            media_type="application/octet-stream",
            filename="oiltech.db"
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not download database: {str(e)}")

@app.delete("/admin/clear-quotations")
def clear_quotations(db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    """
    Elimina solamente las cotizaciones/ventas (tabla quotations).
    No toca productos ni usuarios.
    """
    try:
        deleted = db.query(models.Quotation).delete(synchronize_session=False)
        db.commit()
        return {"ok": True, "deleted": deleted}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Could not clear quotations: {str(e)}")

@app.put("/quotations/{quotation_id}/items")
def update_quotation_items(quotation_id: int, items: List[schemas.QuotationItem], db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    db_quotation = db.query(models.Quotation).filter(models.Quotation.id == quotation_id).first()
    if db_quotation is None:
        raise HTTPException(status_code=404, detail="Quotation not found")
    
    # Update items
    # Convert Pydantic models to dicts for JSON storage
    items_json = [item.dict() for item in items]
    db_quotation.items = items_json
    
    db.commit()
    return {"ok": True}


@app.put("/admin/change-password")
async def change_password(password_data: schemas.UserCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    """
    Cambiar la contraseña del usuario actual.
    """
    if not password_data.password:
        raise HTTPException(status_code=400, detail="Contraseña no puede estar vacía")
    
    current_user.hashed_password = auth.get_password_hash(password_data.password)
    db.commit()
    return {"message": "Contraseña actualizada exitosamente"}


# Site Assets Endpoints
SITE_IMAGES_DIR = os.path.join(database.VOLUME_PATH, "site-images")
os.makedirs(SITE_IMAGES_DIR, exist_ok=True)

@app.get("/admin/site-assets", response_model=List[schemas.SiteAsset])
def get_site_assets(db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    return db.query(models.SiteAsset).all()

@app.post("/admin/site-assets/{key}")
async def upload_site_asset(
    key: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    asset = db.query(models.SiteAsset).filter(models.SiteAsset.key == key).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset key not found")
    
    content = await file.read()
    # Basic check for images
    if not file.content_type.startswith("image/"):
         raise HTTPException(status_code=400, detail="El archivo debe ser una imagen")

    # Save to volume
    ext = os.path.splitext(file.filename)[1]
    if not ext:
        ext = ".png" # fallback
    
    filename = f"{key}{ext}"
    filepath = os.path.join(SITE_IMAGES_DIR, filename)
    
    with open(filepath, "wb") as f:
        f.write(content)
    
    # Update DB URL
    asset.image_url = f"/api/files/site-images/{filename}"
    db.commit()
    db.refresh(asset)

    return asset

VALID_SITE_ASSET_DISPLAY_MODES = {"single", "carousel"}

@app.put("/admin/site-assets/{key}/mode")
def update_site_asset_mode(
    key: str,
    payload: schemas.SiteAssetModeUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    asset = db.query(models.SiteAsset).filter(models.SiteAsset.key == key).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset key not found")

    if payload.mode not in VALID_SITE_ASSET_DISPLAY_MODES:
        raise HTTPException(status_code=400, detail="Modo inválido: debe ser 'single' o 'carousel'")

    asset.display_mode = payload.mode
    db.commit()
    db.refresh(asset)
    return asset

@app.post("/admin/site-assets/{key}/gallery")
async def add_site_asset_gallery_image(
    key: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    asset = db.query(models.SiteAsset).filter(models.SiteAsset.key == key).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset key not found")

    content = await file.read()
    if not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="El archivo debe ser una imagen")

    ext = os.path.splitext(file.filename)[1]
    if not ext:
        ext = ".png"

    # Nombre único por imagen: a diferencia de la portada, varias conviven a la vez
    filename = f"{key}_{uuid.uuid4().hex[:8]}{ext}"
    filepath = os.path.join(SITE_IMAGES_DIR, filename)

    with open(filepath, "wb") as f:
        f.write(content)

    gallery = list(asset.gallery_urls or [])
    gallery.append(f"/api/files/site-images/{filename}")
    asset.gallery_urls = gallery
    db.commit()
    db.refresh(asset)
    return asset

@app.put("/admin/site-assets/{key}/gallery")
def update_site_asset_gallery(
    key: str,
    payload: schemas.SiteAssetGalleryUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    asset = db.query(models.SiteAsset).filter(models.SiteAsset.key == key).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset key not found")

    if not payload.images:
        raise HTTPException(status_code=400, detail="La lista de imágenes no puede estar vacía")

    known_urls = {asset.image_url, *(asset.gallery_urls or [])}
    if any(url not in known_urls for url in payload.images):
        raise HTTPException(status_code=400, detail="La lista contiene imágenes que no pertenecen a este asset")

    # La posición 0 pasa a ser la portada (image_url); el resto, la galería —
    # así "reordenar", "eliminar" y "promover a portada" son la misma operación.
    asset.image_url = payload.images[0]
    asset.gallery_urls = payload.images[1:]
    db.commit()
    db.refresh(asset)
    return asset

@app.get("/files/site-images/{filename}")
async def serve_site_image(filename: str):
    filepath = os.path.join(SITE_IMAGES_DIR, filename)
    if not os.path.isfile(filepath):
        # Fallback to default imagenes folder if it exists in the app package?
        # Better: just return 404 if not found in volume
        raise HTTPException(status_code=404, detail="Imagen no encontrada")
    
    return FileResponse(filepath)

# Public endpoint to get site assets mapping
@app.get("/site-assets-map")
def get_site_assets_map(db: Session = Depends(get_db)):
    assets = db.query(models.SiteAsset).all()
    result = {}
    for a in assets:
        images = [url for url in [a.image_url, *(a.gallery_urls or [])] if url]
        result[a.key] = {"mode": a.display_mode, "images": images}
    return result
